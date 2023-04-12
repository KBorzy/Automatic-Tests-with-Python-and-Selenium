import openpyxl

class Data:
    @staticmethod
    def get_test_input_data(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        email = sheet.cell(row=1, column=1).value
        password = sheet.cell(row=2, column=1).value

        return {'email': email, 'password': password}
